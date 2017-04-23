import os
import sys
import datetime as dt 
import pandas as pd 
import numpy as np
from openpyxl import load_workbook

#def report_non_unique_file(files_name):
#    counts = pd.Series(files_name).value_counts()
#    counts = counts.tolist()
#    non_uniqe_files = []
#    for file_name , count in files_name, counts:
#        if count > 1:
#            non_uniqe_files.append(file_name)
    
#    if non_uniqe_files != []:
#        error_test = open('Error messsage',"a+")
#        error_test.write("Here are duplicate files in folder %s! \n\n Please check the file name and keep the lastest file in the folder"% non_uniqe_files)
#        error_test.close()  
#        sys.exit()

today = dt.datetime.today().strftime("%Y-%m-%d %I'%M %p")
input_direc = '\\\sinppf301\APL\APL Marketing\Sell sheets\OCEAN ALLIANCE\Latest_change-request-forms\\'
output_direc = '\\\sinppf301\APL\APL Marketing\Sell sheets\OCEAN ALLIANCE\OA Service req consolidation\\'
reference_direc = '\\\sinppf301\APL\APL Marketing\Sell sheets\OCEAN ALLIANCE\OA Service req consolidation\OA Service req reference\\'
ports_file_path = ''.join([reference_direc, 'Ports File 15 Mar 2017.xlsx'])
match_table_path = ''.join([reference_direc, 'Ports Name Matching table.xlsx'])
master_file_path = ''.join([output_direc, "Master_file %s.xlsx"%today])
error_message = ''.join([output_direc, "Error Message %s.txt"%today])
warning_message = ''.join([output_direc,"Warning Message %s.txt"% today])
status_message = ''.join([output_direc,"Status Message %s.txt"% today])
ports_file = pd.read_excel(ports_file_path)
UNCode_table = ports_file[['UNCode','Port' ]]
col_name = ['Alliance/others', 'Alliance/Operator', 'Service code', 'POL','POD','Transit Time (days)', 'Trade','Direction' ]

mincol = 2 # column B
key1 = '10. transit times'
key1_max_row = 2000
direction = ' westbound  eastbound  southbound  northbound '
max_space = 20
max_table_width = 100
max_table_length  =100



## The purpose of funtion is to find out the starting row number of first transit time table.
## Summary of working flow: 1. Seaching key value: '10. transit times' from row 1 to 2000 -> Searching direction words = [' westbound  eastbound  southbound  northbound ']
##                          2. Find the cell where the nearby right and downstair cells are both not None(not empty), recoard the row number of this cell as the starting poiont of the 1st transit table.     
##                          3. Create essor message in master_file folder if we can't direction words or transit time table. 
## --------------------------
def table_minrow1(worksheet):

    i = 1
    while i < key1_max_row:
        i +=1
        if worksheet["B%d"%i].value != None and type(worksheet["B%d"%i].value) != long and key1 in worksheet["B%d"%i].value.lower():
            m = i ## row number of '10. Transit Time' 
            
            for k in range(m, m+ max_space):
                if worksheet["B%d"%k].value != None and worksheet["B%d"%k].value.lower() in direction:
                    direction_row = k ## key row 
                    
                    for l in range(direction_row, direction_row+max_space): ## max 10 rows between the key and it's table
                        start_row_check = (worksheet["C%d"%l].value != None) and (worksheet["B%d"%(l+1)].value != None) 
                        if start_row_check:
                            minrow = l ## left_up corner row number
 
                            return minrow, direction_row
                            break 
                   
                    error_test = open(error_message,"a+")
                    error_test.write("Can not find Transite Table in '%s'!\n"% file_name )
                    error_test.close()
                    break
                
            error_test = open(error_message,"a+")
            error_test.write("Can not find direction in '%s'!\n"% file_name )
            error_test.close()
            break
            

## Purpose: Find out the starting row number of 2nd transit time table.          
## Working Flow: 1. Based on the ending row number of 1st table , Seaching the cell contains valus in [' westbound  eastbound  southbound  northbound ']
##               2. Find the cell where the nearby right and downstair cells are both not None(not empty), recoard the row number of this cell as the starting poiont of the 2nd transit table.
##               3. Create error message it can't find second table and warning message if can'f find cells has word in in [' westbound  eastbound  southbound  northbou
## ----------------

def table_minrow2(worksheet, maxrow):
    
    i = 0
    while i < max_space:
        i +=1
        direction_row = maxrow + i
        if worksheet["B%d"% direction_row].value != None and worksheet["B%d"%direction_row].value.lower() in direction:
                   
            for l in range(direction_row, direction_row+ max_space ): ## max 5 rows between the key and it's table
                start_row_check = (worksheet["C%d"%l].value != None) and (worksheet["B%d"%(l+1)].value != None) 
                if start_row_check:
                    minrow = l ## left_up corner row number
                    return minrow, direction_row
                    break        

            error_test = open(error_message,"a+")
            error_test.write("Find two key words, but there is only one TRANSIT TIMES table in '%s'!\n"% file_name)
            error_test.close()      
            break 
                   

    minrow = None
    direction_row = None
    error_test = open(warning_message,"a+")
    error_test.write("Is it single direction in '%s'?\n"% file_name)
    error_test.close()
    
    return minrow,direction_row



##Purpose: Find out the ending row number of each table, given by the starting row number of it
## Work Flow: 1. Given the starting row number of table
##           (2). Serching the cell where value of it is not long(not number) and its valus is None(empty) or in the direction or it is 'instruction'
##            3. Record the ending row number as the row number of cell in (2) - 1
## -------------
def table_maxrow(worksheet, minrow):
    m = 0 ##  table's row number
    while m < max_table_length:
        m = m +1 
        value = worksheet["B%d"%(minrow+m)].value
        if  type(value) != long and (value == None or value.lower() in direction or 'instruction' in value.lower()):
            max_row = minrow + m-1 ## max_row -- left_down coner row number 
            return max_row     
            break
            


## Purpose: Find out the ending columns of each table
## Work flow: 1. Find out the cell which the right cell near it is None in the last row of the table.
##            2. max_col = The column of cell - 1
## ------------
def table_maxcol(worksheet, maxrow):
    
    k = 1 ## excape first column 
    while k < max_table_width :
        k +=1 
        if worksheet.cell(row = maxrow, column = k).value == None: 
            max_col = k-1 ## k -- right_down coner column number
            return max_col     
            break
        

## Purpose: Select the range of transit table we need and change the type of table for further transformation
    
def table_range(worksheet,minrow,maxrow,maxcol):
    
    data = []
    length = maxcol - mincol +1 
    width = maxrow - minrow +1 
    table_range = worksheet.iter_cols(min_row=minrow, min_col = mincol , max_row= maxrow, max_col= maxcol)
    for row in table_range:
        for cell in row:
            data.append(cell.value)
    data = np.array(data)
    data = data.reshape((length,width))
    return data
 

## Purpose: Remove the unwanted characters in the table if the table type is normal type( The origin ports are listed at column B )
## --------
def normal_type(data):
    
    data = np.swapaxes(data, 0, 1)
    col_name = data[0,1:]
    index = data[1:,0]
    data_in = data[1:,1:]
    
    for i in range(0, len(col_name)):
        col_name[i] = col_name[i].encode('ascii', 'ignore')
        
    for j in range(0,len(index)):
        index[j] = index[j].encode('ascii', 'ignore')
    
    return data_in, col_name,index

## Purpose: Remove the unwanted characters in the table if the table type is reverse type( The origin ports are listed at end column of table )
## --------
def reverse_type(data):
    index  = data[:-1,0]
    col_name = data[-1,1:]
    data_in = data[:-1, 1:]
    
    for i in range(0, len(col_name)):
        col_name[i] = col_name[i].encode('ascii', 'ignore')
    for j in range(0,len(index)):
        index[j] = index[j].encode('ascii', 'ignore')
    
    return data_in, col_name,index

## Purpose : Detect the type of transite tiem table ( Normal type/ reverse type) and return to the respective funtion above. 
## ---------
def return_matrix(worksheet,data,minrow):
    if type(worksheet['B%d'% (minrow+1)].value) == unicode :
        return normal_type(data)
    else: 
        return reverse_type(data)
                   

## Purpose: Conver the original table to the table format needed in master file; and Read message from file name into the other columns
## Work flow: 1. stack() funtion can help to achieve new format need in master file
##            2. Add other columns and fill in their value from file name
## ------------
def final_table(data_in, col_name,index,file_name, direction):
    
    table_df = pd.DataFrame(data = data_in, columns = col_name, index = index)
    df = table_df.stack().reset_index().rename(columns={'level_0':'Departure','level_1':'Arrive', 0:'Transit Time (days)'})
    
    trade_name = file_name.split('_')[0]
    sve_name =  ' - '.join([file_name.split('_')[3],file_name.split('_')[2]])
    alliance = file_name.split('_')[1]
    
    df['Alliance/others']= alliance
    df['Alliance/Operator'] = alliance
    df['Trade'] = trade_name
    df['Service code'] = sve_name
    df['Direction'] = direction
    
    return df


## Purpose: Union above functions. Start from reading files from folder to covert them to master file format. 
## 
def _main_(files_name):
    
    database = {}
    for file_name in files_name:


        workbook = load_workbook(filename = file_name, data_only=True)
        worksheet = workbook[workbook.sheetnames[-1]]
        ## [-1] => Reading the last sheet in the file 
        name = file_name.replace('.xlsx', '')
        try:
        ## Try - Except : Try to escape the error and move on to concatenating each files
            minrow1,direction_row1 = table_minrow1(worksheet)
            maxrow1 = table_maxrow(worksheet, minrow1)
            maxcol1 = table_maxcol(worksheet, maxrow1)
            direction1 = worksheet.cell(row = direction_row1, column = mincol).value
            data1 = table_range(worksheet,minrow1,maxrow1,maxcol1)
            data_in1, col_name1, index1 = return_matrix(worksheet,data1,minrow1)
            table1 = final_table(data_in1, col_name1,index1,name, direction1)
    
            minrow2,direction_row2 = table_minrow2(worksheet,maxrow1)
            if direction_row2 != None: ## if the second direction exist in the file, which means it contains two table(two directions) 
                direction2 = worksheet.cell(row = direction_row2, column =2).value
                maxrow2 = table_maxrow(worksheet, minrow2)
                maxcol2 = table_maxcol(worksheet, maxrow2) 
                data2 = table_range(worksheet,minrow2,maxrow2,maxcol2)
                data_in2, col_name2, index2 = return_matrix(worksheet,data2,minrow2)
                table2 = final_table(data_in2, col_name2,index2,name, direction2)
                file_table = pd.concat([table1, table2]).reset_index().drop('index', axis =1)
            else: ## There is only one table exist in the file 
                file_table = table1 
            
            database[name] = file_table ## This database stores the information structure like this: [ file1: file1_table, file2: file2_table, file3: file3_table,....] 
        except:
            pass

    df_values = database.values()
    master_table = pd.concat(df_values, ignore_index = True)
    return master_table, database

## The final funtion to transfer every request file in 5 folders and combine them to the master file.   
def request_form_converter():
    data = {}
    data_dic = {} ## It helps us to record the data from every files and folder. We gonna use this variable to creat status message
    for folder in os.listdir(input_direc):  ## go through each trade folder 
        folder_path = ''.join([input_direc, folder]) 
        files_name = os.listdir(folder_path) ## go through each file in the certain folder. 
        files_xlsx = [f for f in files_name if f[-4:]== 'xlsx' and f[:2] != '~$'] ## filter the files end with 'xlsx' and not start with '~$' 
        os.chdir(folder_path)
        master_table, database = _main_(files_xlsx) 
        data[folder] = master_table
        data_dic[folder] = database 
        
    master_values = data.values()
    master_tables = pd.concat(master_values, ignore_index = True)

    return master_tables, data_dic 
## ---------------------------------------------------------------------
## Step 2:

## clean function => strip blanks at the beginning/end and lowercase the words  
def clean_port_name(df_Series):
    return df_Series.str.strip().str.lower()

## Read from 'Ports name matching table' and replace the ports name from master file to the corresponding name in Ports file.
def correct_non_matching_port_name(master_file):
    matching_table = pd.read_excel(match_table_path)
    dic_match = matching_table.set_index('Ports Name in Master file')['Ports Name in Ports file'].to_dict() 
    master_file['Arrive'] = clean_port_name(master_file.Arrive)
    master_file['Departure'] = clean_port_name(master_file.Departure)
    for k, v in dic_match.iteritems():
        master_file['Departure'] = master_file['Departure'].replace(k,v)
        master_file['Arrive'] = master_file['Arrive'].replace(k,v)
    return master_file

## clean funtion and add two UnCode columns 
def master_file_clean(master_file):
    master_file['Arrive'] = clean_port_name(master_file.Arrive)
    master_file['Departure'] = clean_port_name(master_file.Departure)
    UNCode_table['Port'] = clean_port_name(UNCode_table.Port)
    master_file['POL'] = ''
    master_file['POD'] =''
    return master_file

## list all unique port names from Departure & Arrive
## Uncode1 = "Uncode_Departure" & Uncode2 = "Uncode_Arrive"
## Departure = "Departure" & Arrive = "Arrive"

def port_name_to_uncode(UNCode_table,master_file, Departure, Arrive,Uncode1, Uncode2):
    
    ports_name = UNCode_table.Port.tolist()
    ports = {}
    non_match_departure = []
    non_match_arrive = []
    
## ports is a dictionary, key: port name, value: UNCode
    for port in ports_name:
        ports[port] = UNCode_table.UNCode[UNCode_table.Port == port].tolist()
    
    master_ports_departure = master_file[Departure].unique().tolist()
    master_ports_arrive = master_file[Arrive].unique().tolist()
    
    for name in master_ports_departure:
        if name in ports.keys():
            master_file.loc[master_file[Departure] == name, 'POL'] = ports[name][0]
        else:
            non_match_departure.append(name)
            
    for name in master_ports_arrive:
        if name in ports.keys():
            master_file.loc[master_file[Arrive] == name, 'POD'] = ports[name][0]
        else:
            non_match_arrive.append(name)
    
    non_match = np.unique(non_match_arrive + non_match_departure).tolist()
    if non_match != []:
        non_match = [str(name) for name in non_match ]   
        match_error = open(error_message,"a+")
        match_error.write("\nCan not find such port names %s in the Ports File!\n\n\nPlease correct the names in excel: 'Ports Name Matching table.xlsx'"% non_match)
        match_error.close()
        table = pd.DataFrame({'Ports Name in Master file':non_match,'Ports Name in Ports file': ''})
        matching_table = pd.read_excel(match_table_path)
        matching_table_final = pd.concat([matching_table,table], ignore_index = True)
        matching_table_final.to_excel(match_table_path, index = False)
    else:
        
        master_file = master_file.drop(['Departure', 'Arrive'], axis =1)
        master_file=master_file[col_name]
        master_file = master_file[master_file['Transit Time (days)'] != '-']
        master_file.to_excel(master_file_path, index = False)  

def status_report(data_dic):
    name = []
    warning_name = []
    for f in data_dic.keys():
        for file_key in data_dic[f].keys():
            name = name + [file_key]
            value = data_dic[f][file_key]
            if len(value) == 0:
                warning_name.append(file_key)          

    if len(warning_name) != 0:
        error = open(error_message,"a+")
        match_error.write("\n For this run, we failed to transfer these files %s to the master file! \n\n\n Please check them!"% warning_name)
        match_error.close()

    else:
        name = pd.DataFrame(name)[0]
        pd.set_option('display.max_rows', None)
        pd.set_option('display.max_columns', None)
        status = open(status_message,"a+")
        status.write("\nFor this run, we have successfully read %s request files from the 'Latest_change-request-forms' folder. Here is the list of files we read: "%len(name))
        status.write("\n\n{} \n\nPlease check them with the actual files in folder: Latest_change-request-forms.".format(name))
        status.close()
        port_name_to_uncode(UNCode_table,master_file,"Departure", "Arrive","POL" ,"POD")

        
master_file, data_dic = request_form_converter()
master_file = correct_non_matching_port_name(master_file)
master_file = master_file_clean(master_file)
status_report(data_dic)
